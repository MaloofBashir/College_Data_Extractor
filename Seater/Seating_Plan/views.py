from datetime import datetime
from io import BytesIO

from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect, render

from .result_parser import summarize_result_pdf
from .attendance_pdf import generate_attendance_pdf
from .seating import build_attendance_summary, build_filtered_attendance_summary, getting_all_subs, json_data, merge_rolls, sorted_centres
from .summary_pdf import generate_summary_pdf
from openpyxl import Workbook


merged_data={}
new_dict={}

center_nos=[]
filtered_data={}
current_data_of_centre=[]

filtered=False


def parse_subject_filters(request):
    raw_value = request.GET.get("subjects", "")
    return [item.strip() for item in raw_value.split(",") if item.strip()]


def parse_centre_filter(request, attendance_data=None):
    selected_centre = request.GET.get("centre", "").strip()
    if attendance_data:
        centres = sorted_centres(attendance_data)
        if selected_centre in attendance_data:
            return selected_centre
        return centres[0] if centres else ""
    return selected_centre


def attendance_summary_payload(data, selected_centre=None, selected_subjects=None):
    summary = build_filtered_attendance_summary(data, selected_centre=selected_centre, selected_subjects=selected_subjects)
    return {
        "selected_centre": summary["selected_centre"],
        "centres": summary["centres"],
        "available_subjects": [item["subject"] for item in summary["subjects"]],
        "selected_subjects": summary["selected_subjects"],
        "subjects": summary["filtered_subjects"],
        "selected_rolls": summary["selected_rolls"],
        "subject_count": summary["subject_count"],
        "selected_roll_count": summary["selected_roll_count"],
    }
# Create your views here.
def index(request):
    return render(request, 'index.html')


def result_summary(request):
    context = {
        "current_date_time": datetime.now(),
    }

    if request.method == 'POST' and 'file' in request.FILES:
        uploaded_file = request.FILES['file']
        try:
            context["summary"] = summarize_result_pdf(uploaded_file)
            request.session["result_summary"] = context["summary"]
            request.session["summary_generated_at"] = datetime.now().strftime("%d %b %Y, %I:%M %p")
            context["summary_generated_at"] = request.session["summary_generated_at"]
        except Exception:
            context["error"] = (
                "Error in uploading the result PDF. Please ensure you upload a KU result PDF in the expected format."
            )
        return render(request, 'result_summary.html', context)

    if request.session.get("result_summary"):
        context["summary"] = request.session.get("result_summary")
        context["summary_generated_at"] = request.session.get("summary_generated_at")

    return render(request, 'result_summary.html', context)


def download_result_summary_pdf(request):
    summary = request.session.get("result_summary")
    if not summary:
        return redirect("result_summary")

    pdf_bytes = generate_summary_pdf(summary)
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="result-summary.pdf"'
    return response


def download_attendance_pdf(request):
    attendance_data = request.session.get("attendance_data")
    if not attendance_data:
        return redirect("Table_rollno")

    selected_centre = parse_centre_filter(request, attendance_data)
    selected_subjects = parse_subject_filters(request)
    summary = build_filtered_attendance_summary(attendance_data, selected_centre=selected_centre, selected_subjects=selected_subjects)
    pdf_bytes = generate_attendance_pdf(summary)
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="attendance-rolls-centre-{selected_centre or "sheet"}.pdf"'
    return response

def Table_rollno(request):
    
    global new_dict,filtered_data,center_nos
    context = {}
    center_nos.clear()
    merged_data={}
    date_time=datetime.now()
    context["current_date_time"]=date_time

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        if not request.session.get("attendance_data"):
            return JsonResponse({"error": "No attendance data found."}, status=400)

        selected_centre = parse_centre_filter(request, request.session["attendance_data"])
        selected_subjects = parse_subject_filters(request)
        payload = attendance_summary_payload(
            request.session["attendance_data"],
            selected_centre=selected_centre,
            selected_subjects=selected_subjects,
        )
        filtered_data.clear()
        filtered_data.update({item["subject"]: item["roll_numbers"] for item in payload["subjects"]})
        return JsonResponse(payload, safe=False)
    
    if request.method == 'POST' and 'file' in request.FILES:
        filtered=False
        filtered_data.clear()
        merged_data.clear()
        new_dict.clear()

        
        center_nos.clear()
        uploaded_file = request.FILES['file'] 
        try:
           
            getting_all_subs(uploaded_file)
        except Exception:
            error_message = "Error in uploading the file, Please ensure you upload a Pdf-Attendance Sheet from KU"
            context["error"] = error_message
        new_dict.update(json_data.copy())
        merged_data=merge_rolls(new_dict)
        filtered_data.update(merged_data.copy())
        default_centre = sorted_centres(new_dict)[0] if new_dict else ""
        summary = build_filtered_attendance_summary(new_dict, selected_centre=default_centre)
        request.session["attendance_data"] = new_dict
        request.session["attendance_generated_at"] = datetime.now().strftime("%d %b %Y, %I:%M %p")

        context["attendance_summary"] = summary
        context["attendance_generated_at"] = request.session["attendance_generated_at"]
        context["all_subject_names"] = [item["subject"] for item in summary["subjects"]]
        context["centre_nos"] = sorted_centres(new_dict)
        json_data.clear()
        response = render(request, 'Table_rollno.html', context)
        return response
    else:
        if request.session.get("attendance_data"):
            default_centre = sorted_centres(request.session["attendance_data"])[0] if request.session["attendance_data"] else ""
            summary = build_filtered_attendance_summary(request.session["attendance_data"], selected_centre=default_centre)
            context["attendance_summary"] = summary
            context["attendance_generated_at"] = request.session.get("attendance_generated_at")
            context["all_subject_names"] = [item["subject"] for item in summary["subjects"]]
            context["centre_nos"] = sorted_centres(request.session["attendance_data"])
        return render(request, 'Table_rollno.html', context)

def export_excel(request):
    attendance_data = request.session.get("attendance_data")
    if not attendance_data:
        return redirect("Table_rollno")

    selected_centre = parse_centre_filter(request, attendance_data)
    selected_subjects = parse_subject_filters(request)
    summary = build_filtered_attendance_summary(
        attendance_data,
        selected_centre=selected_centre,
        selected_subjects=selected_subjects,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = selected_centre or "Sheet1"

    for col, item in enumerate(summary["filtered_subjects"], start=1):
        ws.cell(row=1, column=col, value=item["subject"])
        for row, value in enumerate(item["roll_numbers"], start=2):
            ws.cell(row=row, column=col, value=value)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    file_name = f'roll_numbers_{selected_centre or "all"}.xlsx'
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{file_name}"'
    return response
