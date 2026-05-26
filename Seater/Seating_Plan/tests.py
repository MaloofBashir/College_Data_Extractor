from io import BytesIO
from unittest.mock import patch

from openpyxl import load_workbook
from django.test import Client, SimpleTestCase, TestCase
from django.urls import reverse

from .result_parser import parse_overall_result, parse_subject_grades, split_student_blocks, summarize_result_pdf
from .seating import build_filtered_attendance_summary


class ResultParserTests(SimpleTestCase):
    def test_split_student_blocks_handles_pdf_layout(self):
        page_text = """
        Roll No.Reg No. Name Subjects ResultGovt. Degree College
        25530344 559-zp-2022 FAAZIL MANZOOR HST522J1(58;20) HST522J2(36;42)
        HST522J3(55;42) URL522N(54;25) <br>HST522J1(B+) HST522J2(C+)
        HST522J3(B) URL522N(B+) <br>PASS/ SGPA--> 7.1
        25535774 453-ZP-2022 SAIQA MAJEED PLS522N(35;16) EDU522J1(27;18)
        EDU522J2(52;28) EDU522J3(23;27) <br>PLS522N(C+) EDU522J1(C)
        EDU522J2(C+) EDU522J3(F) <br>R EDU522J3
        """

        blocks = split_student_blocks(page_text)

        self.assertEqual(len(blocks), 2)

    def test_parse_subject_grades_reads_grade_section(self):
        block = (
            "25535774 453-ZP-2022 SAIQA MAJEED PLS522N(35;16) EDU522J1(27;18) "
            "EDU522J2(52;28) EDU522J3(23;27) <br>PLS522N(C+) EDU522J1(C) "
            "EDU522J2(C+) EDU522J3(F) <br>R EDU522J3"
        )

        parsed = parse_subject_grades(block)

        self.assertEqual(parsed["PLS522N"], "PASS")
        self.assertEqual(parsed["EDU522J3"], "FAIL")

    def test_parse_overall_result_uses_pass_sgpa_marker(self):
        block = "ABC <br>SUB1(B+) SUB2(C) <br>PASS/ SGPA--> 7.1"

        parsed = parse_overall_result(block)

        self.assertEqual(parsed, "PASS")

    def test_parse_overall_result_detects_reappear(self):
        block = "ABC <br>SUB1(F) SUB2(C) <br>R SUB1"

        parsed = parse_overall_result(block)

        self.assertEqual(parsed, "REAPPEAR")

    @patch("Seating_Plan.result_parser.PdfReader")
    def test_summarize_result_pdf_builds_status_and_subject_totals(self, pdf_reader_mock):
        page_text = """
        Roll No.Reg No. Name Subjects ResultGovt. Degree College
        25530344 559-zp-2022 FAAZIL MANZOOR HST522J1(58;20) HST522J2(36;42)
        HST522J3(55;42) URL522N(54;25) <br>HST522J1(B+) HST522J2(C+)
        HST522J3(B) URL522N(B+) <br>PASS/ SGPA--> 7.1
        25535774 453-ZP-2022 SAIQA MAJEED PLS522N(35;16) EDU522J1(27;18)
        EDU522J2(52;28) EDU522J3(23;27) <br>PLS522N(C+) EDU522J1(C)
        EDU522J2(C+) EDU522J3(F) <br>R EDU522J3
        25535775 454-ZP-2022 SAMPLE STUDENT URL522N(0;0) <br>URL522N(AB) <br>FAIL
        """

        pdf_reader_mock.return_value.pages = [type("Page", (), {"extract_text": lambda self: page_text})()]

        summary = summarize_result_pdf(object())

        self.assertEqual(summary["total_students"], 3)
        self.assertEqual(summary["overall_passed"], 1)
        self.assertEqual(summary["overall_reappear"], 1)
        self.assertEqual(summary["overall_failed"], 1)
        self.assertEqual(summary["average_sgpa_pass_only"], 7.1)
        self.assertEqual(summary["subjects"][0]["subject_name"], "Education")
        self.assertEqual(summary["students"][1]["issue_subjects"], ["EDU522J3"])


class AttendanceSummaryTests(TestCase):
    def test_build_filtered_attendance_summary_filters_multiple_subjects(self):
        attendance_data = {
            "101": {
                "MAT101": ["1001", "1002"],
                "PHY101": ["1002", "1003"],
            },
            "102": {
                "MAT101": ["1004"],
                "CHE101": ["1005"],
            },
        }

        summary = build_filtered_attendance_summary(
            attendance_data,
            selected_centre="102",
            selected_subjects=["MAT101", "CHE101"],
        )

        self.assertEqual(summary["selected_subject_count"], 2)
        self.assertEqual(summary["selected_roll_count"], 2)
        self.assertEqual(summary["selected_rolls"], ["1004", "1005"])
        self.assertEqual([item["subject"] for item in summary["filtered_subjects"]], ["CHE101", "MAT101"])

    def test_build_filtered_attendance_summary_filters_by_centre_first(self):
        attendance_data = {
            "101": {
                "MAT101": ["1001", "1002"],
                "PHY101": ["1002", "1003"],
            },
            "102": {
                "MAT101": ["2001"],
                "CHE101": ["2002"],
            },
        }

        summary = build_filtered_attendance_summary(attendance_data, selected_centre="102", selected_subjects=["MAT101"])

        self.assertEqual(summary["selected_centre"], "102")
        self.assertEqual(summary["selected_rolls"], ["2001"])
        self.assertEqual([item["subject"] for item in summary["subjects"]], ["CHE101", "MAT101"])

    def test_export_excel_streams_filtered_workbook_without_media_directory(self):
        client = Client()
        session = client.session
        session["attendance_data"] = {
            "101": {
                "MAT101": ["1001", "1002"],
                "PHY101": ["1002", "1003"],
            },
            "102": {
                "MAT101": ["2001"],
                "CHE101": ["2002"],
            },
        }
        session.save()

        response = client.get(reverse("export_excel"), {"centre": "102", "subjects": "CHE101,MAT101"})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn('filename="roll_numbers_102.xlsx"', response["Content-Disposition"])

        workbook = load_workbook(filename=BytesIO(response.content))
        sheet = workbook.active

        self.assertEqual(sheet.title, "102")
        self.assertEqual(sheet.cell(row=1, column=1).value, "CHE101")
        self.assertEqual(sheet.cell(row=2, column=1).value, "2002")
        self.assertEqual(sheet.cell(row=1, column=2).value, "MAT101")
        self.assertEqual(sheet.cell(row=2, column=2).value, "2001")

# Create your tests here.
