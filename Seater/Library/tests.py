from io import BytesIO

from django.http import Http404
from django.test import RequestFactory, TestCase
from django.urls import reverse
from openpyxl import Workbook, load_workbook

from .forms import BookForm, EmployeeRegistrationForm
from .models import Book, Employee
from .views import admin_dashboard, all_books, dashboard, delete_book, delete_employee, filter_books, landing


class LibraryWorkflowTests(TestCase):
    def setUp(self):
        self.factory = RequestFactory()
        self.employee = Employee.objects.create(
            email="staff@example.com",
            first_name="Library",
            last_name="Staff",
        )
        self.employee.set_password("StrongPass123!")
        self.employee.save(update_fields=["password_hash"])

    def login(self, employee=None):
        session = self.client.session
        session["library_employee_id"] = (employee or self.employee).id
        session.save()

    def book_payload(self, **overrides):
        payload = {
            "title": "Introduction to History",
            "author": "A. Writer",
            "accession_number": "ACC-1001",
            "isbn": "HIST-101",
            "locker_no": "L-12",
            "subject": "History",
            "quantity": 3,
            "publisher": "College Press",
            "remarks": "Core reading",
        }
        payload.update(overrides)
        return payload

    def test_register_logs_employee_in_and_normalizes_email(self):
        response = self.client.post(reverse("library_register"), {
            "email": "NEW.STAFF@Example.COM",
            "first_name": "New",
            "last_name": "Staff",
            "password": "StrongPass123!",
        })

        employee = Employee.objects.get(email="new.staff@example.com")
        self.assertRedirects(response, reverse("library_dashboard"), fetch_redirect_response=False)
        self.assertEqual(self.client.session["library_employee_id"], employee.id)
        self.assertTrue(employee.check_password("StrongPass123!"))

    def test_register_rejects_case_insensitive_duplicate_email(self):
        form = EmployeeRegistrationForm({
            "email": "STAFF@EXAMPLE.COM",
            "first_name": "Another",
            "last_name": "Employee",
            "password": "StrongPass123!",
        })

        self.assertFalse(form.is_valid())
        self.assertIn("already registered", form.errors["email"][0])
        self.assertEqual(Employee.objects.count(), 1)

    def test_public_books_page_and_login_returns_to_requested_page(self):
        Book.objects.create(added_by=self.employee, **self.book_payload())
        Book.objects.create(
            added_by=self.employee,
            **self.book_payload(title="Political Science", quantity=5, subject="Political Science"),
        )
        request = self.factory.get(reverse("library_all_books"))
        request.session = {}
        response = all_books(request)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Total book records")
        self.assertContains(response, "<strong>2</strong>", html=True)

        request = self.factory.get(reverse("library_all_books"), {"q": "History"})
        request.session = {}
        response = all_books(request)

        self.assertContains(response, "Introduction to History")
        self.assertContains(response, "Political Science")

        response = self.client.post(reverse("library_login"), {
            "email": "STAFF@EXAMPLE.COM",
            "password": "StrongPass123!",
            "next": reverse("library_all_books"),
        })

        self.assertRedirects(response, reverse("library_all_books"), fetch_redirect_response=False)

    def test_login_does_not_redirect_to_external_url(self):
        response = self.client.post(reverse("library_login"), {
            "email": self.employee.email,
            "password": "StrongPass123!",
            "next": "https://example.net/untrusted",
        })

        self.assertRedirects(response, reverse("library_dashboard"), fetch_redirect_response=False)

    def test_legacy_employee_login_sets_first_password(self):
        legacy = Employee.objects.create(email="legacy@example.com", first_name="Legacy", last_name="Staff")

        response = self.client.post(reverse("library_login"), {
            "email": legacy.email,
            "password": "LegacyPass123!",
        })
        legacy.refresh_from_db()

        self.assertRedirects(response, reverse("library_dashboard"), fetch_redirect_response=False)
        self.assertTrue(legacy.check_password("LegacyPass123!"))

    def test_employee_can_add_update_filter_and_delete_own_book(self):
        self.login()
        response = self.client.post(reverse("library_dashboard"), self.book_payload())
        book = Book.objects.get()

        self.assertRedirects(response, reverse("library_dashboard"), fetch_redirect_response=False)
        self.assertEqual(book.added_by, self.employee)
        self.assertEqual(book.accession_number, "ACC-1001")

        response = self.client.post(
            reverse("library_dashboard"),
            self.book_payload(book_id=book.id, title="Modern History"),
        )
        book.refresh_from_db()

        self.assertRedirects(response, reverse("library_dashboard"), fetch_redirect_response=False)
        self.assertEqual(book.title, "Modern History")
        self.assertEqual(book.remarks, "Core reading")

        request = self.factory.get(reverse("library_all_books"), {"q": "Modern"})
        books, query, subject = filter_books(request)
        self.assertEqual(list(books), [book])
        self.assertEqual((query, subject), ("Modern", ""))

        request = self.factory.get(reverse("library_delete_book", args=[book.id]))
        request.library_employee = self.employee
        request.session = {}
        response = delete_book.__wrapped__(request, book.id)
        self.assertEqual(response.status_code, 200)
        captcha_answer = request.session[f"delete_book_captcha_{book.id}"]
        session = self.client.session
        session[f"delete_book_captcha_{book.id}"] = captcha_answer
        session.save()
        response = self.client.post(reverse("library_delete_book", args=[book.id]), {
            "captcha_answer": captcha_answer,
        })
        self.assertRedirects(response, reverse("library_dashboard"), fetch_redirect_response=False)
        self.assertFalse(Book.objects.exists())

    def test_book_form_omits_removed_fields_and_matches_text_case_insensitively(self):
        Book.objects.create(added_by=self.employee, **self.book_payload())
        form = BookForm(self.book_payload(
            title="introduction to history",
            author="a. writer",
            subject="HISTORY",
        ))

        self.assertNotIn("quantity", form.fields)
        self.assertNotIn("isbn", form.fields)
        self.assertNotIn("publisher", form.fields)
        self.assertTrue(form.is_valid())
        self.assertEqual(form.cleaned_data["title"], "Introduction to History")
        self.assertEqual(form.cleaned_data["author"], "A. Writer")
        self.assertEqual(form.cleaned_data["subject"], "History")

    def test_subject_filter_matches_selected_subject(self):
        history = Book.objects.create(added_by=self.employee, **self.book_payload())
        Book.objects.create(added_by=self.employee, **self.book_payload(title="Other", subject="Urdu"))

        request = self.factory.get(reverse("library_all_books"), {"subject": " history "})
        books, query, subject = filter_books(request)

        self.assertEqual(list(books), [history])
        self.assertEqual((query, subject), ("", "history"))

    def test_employee_cannot_change_or_delete_another_employees_book(self):
        other = Employee.objects.create(email="other@example.com", first_name="Other", last_name="Staff")
        book = Book.objects.create(added_by=other, **self.book_payload())
        update_request = self.factory.post(
            reverse("library_dashboard"),
            self.book_payload(book_id=book.id, title="Changed title"),
        )
        update_request.library_employee = self.employee
        delete_request = self.factory.get(reverse("library_delete_book", args=[book.id]))
        delete_request.library_employee = self.employee

        with self.assertRaises(Http404):
            dashboard.__wrapped__(update_request)
        with self.assertRaises(Http404):
            delete_book.__wrapped__(delete_request, book.id)
        book.refresh_from_db()
        self.assertEqual(book.title, "Introduction to History")

    def test_logged_in_recorded_books_show_only_own_books_without_actions(self):
        own = Book.objects.create(added_by=self.employee, **self.book_payload())
        other = Employee.objects.create(email="other@example.com", first_name="Other", last_name="Staff")
        Book.objects.create(added_by=other, **self.book_payload(title="Other book"))
        request = self.factory.get(reverse("library_all_books"))
        request.session = {"library_employee_id": self.employee.id}

        response = all_books(request)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, own.title)
        self.assertNotContains(response, "Other book")
        self.assertNotContains(response, "Update")
        self.assertNotContains(response, "Delete")

    def test_landing_renders_employee_forms(self):
        request = self.factory.get(reverse("library_home"))
        request.session = {}

        response = landing(request)

        self.assertEqual(response.status_code, 200)
        self.assertIn(b"Register employee", response.content)

    def test_dashboard_renders_book_entry_form(self):
        Book.objects.create(added_by=self.employee, **self.book_payload())
        other = Employee.objects.create(email="other@example.com", first_name="Other", last_name="Staff")
        Book.objects.create(added_by=other, **self.book_payload(title="Other book", quantity=8))
        request = self.factory.get(reverse("library_dashboard"))
        request.library_employee = self.employee

        response = dashboard.__wrapped__(request)

        self.assertEqual(response.status_code, 200)
        self.assertIn(b"Add book", response.content)
        self.assertIn(b'<datalist id="subject-suggestions">', response.content)
        self.assertIn(b'<option value="History">', response.content)
        self.assertIn(b"Your entry totals", response.content)
        self.assertIn(b"Your books", response.content)
        self.assertNotIn(b"All books directory", response.content)
        self.assertNotIn(b"Books by employee", response.content)
        self.assertContains(response, "<strong>1</strong>", count=1, html=True)
        self.assertNotIn(b"ISBN / Book no", response.content)
        self.assertNotIn(b"Qty", response.content)
        self.assertNotIn(b"Total quantity", response.content)
        self.assertNotIn(b'name="quantity"', response.content)
        self.assertNotIn(b'name="isbn"', response.content)
        self.assertNotIn(b'name="publisher"', response.content)

    def test_admin_dashboard_requires_admin_flag(self):
        self.login()

        response = self.client.get(reverse("library_admin_dashboard"))

        self.assertRedirects(response, reverse("library_dashboard"), fetch_redirect_response=False)

    def test_admin_can_view_dashboard_and_download_excel_report(self):
        self.employee.is_admin = True
        self.employee.save(update_fields=["is_admin"])
        self.login()
        Book.objects.create(added_by=self.employee, **self.book_payload())

        request = self.factory.get(reverse("library_admin_dashboard"))
        request.library_employee = self.employee
        dashboard_response = admin_dashboard.__wrapped__(request)
        export_response = self.client.get(reverse("library_export_books_excel"))

        self.assertEqual(dashboard_response.status_code, 200)
        self.assertContains(dashboard_response, "Books by locker")
        self.assertContains(dashboard_response, "Books by subject")
        self.assertEqual(
            export_response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(filename=BytesIO(export_response.content))
        self.assertEqual(workbook.sheetnames, ["All Books", "By Subject", "By Locker", "By Employee"])
        self.assertEqual(workbook["All Books"].cell(row=2, column=1).value, "Introduction to History")
        self.assertEqual(workbook["All Books"].cell(row=2, column=3).value, "ACC-1001")
        self.assertNotIn("Quantity", [cell.value for cell in workbook["All Books"][1]])
        self.assertNotIn("ISBN / Book No", [cell.value for cell in workbook["All Books"][1]])
        self.assertNotIn("Publisher", [cell.value for cell in workbook["All Books"][1]])

    def test_admin_can_download_bulk_upload_format(self):
        self.employee.is_admin = True
        self.employee.save(update_fields=["is_admin"])
        self.login()

        response = self.client.get(reverse("library_book_upload_template"))

        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(filename=BytesIO(response.content))
        self.assertEqual(workbook.sheetnames, ["Books Upload"])
        self.assertEqual(
            [cell.value for cell in workbook["Books Upload"][1]],
            ["Title", "Author", "Accession Number", "Locker No", "Subject", "Remarks"],
        )

    def test_admin_can_bulk_upload_books_from_excel(self):
        self.employee.is_admin = True
        self.employee.save(update_fields=["is_admin"])
        self.login()
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Title", "Author", "Accession Number", "Locker No", "Subject", "Remarks"])
        sheet.append(["Introduction to History", "A. Writer", "ACC-1001", "L-12", "History", "Core reading"])
        sheet.append(["Political Science", "B. Author", "", "L-14", "Political Science", ""])
        upload = BytesIO()
        workbook.save(upload)
        upload.seek(0)
        upload.name = "books.xlsx"

        response = self.client.post(reverse("library_bulk_upload_books"), {"books_file": upload})

        self.assertRedirects(response, reverse("library_admin_dashboard"), fetch_redirect_response=False)
        self.assertEqual(Book.objects.count(), 2)
        self.assertTrue(Book.objects.filter(title="Introduction to History", added_by=self.employee).exists())
        self.assertTrue(Book.objects.filter(title="Political Science", accession_number="").exists())

    def test_bulk_upload_reports_row_errors_without_importing_invalid_rows(self):
        self.employee.is_admin = True
        self.employee.save(update_fields=["is_admin"])
        self.login()
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Title", "Author", "Accession Number", "Locker No", "Subject", "Remarks"])
        sheet.append(["", "A. Writer", "ACC-1001", "L-12", "History", "Core reading"])
        upload = BytesIO()
        workbook.save(upload)
        upload.seek(0)
        upload.name = "books.xlsx"

        response = self.client.post(reverse("library_bulk_upload_books"), {"books_file": upload}, follow=True)

        self.assertEqual(Book.objects.count(), 0)
        self.assertContains(response, "Row 2: missing Title.")

    def test_admin_can_delete_other_employees_book(self):
        self.employee.is_admin = True
        self.employee.save(update_fields=["is_admin"])
        other = Employee.objects.create(email="other@example.com", first_name="Other", last_name="Staff")
        book = Book.objects.create(added_by=other, **self.book_payload())

        request = self.factory.get(reverse("library_delete_book", args=[book.id]))
        request.library_employee = self.employee
        request.session = {}
        response = delete_book.__wrapped__(request, book.id)
        captcha_answer = request.session[f"delete_book_captcha_{book.id}"]

        self.assertEqual(response.status_code, 200)
        session = self.client.session
        session["library_employee_id"] = self.employee.id
        session[f"delete_book_captcha_{book.id}"] = captcha_answer
        session.save()
        response = self.client.post(reverse("library_delete_book", args=[book.id]), {
            "captcha_answer": captcha_answer,
        })

        self.assertRedirects(response, reverse("library_admin_dashboard"), fetch_redirect_response=False)
        self.assertFalse(Book.objects.filter(pk=book.id).exists())

    def test_admin_can_delete_employee_without_deleting_their_books(self):
        self.employee.is_admin = True
        self.employee.save(update_fields=["is_admin"])
        self.login()
        other = Employee.objects.create(email="other@example.com", first_name="Other", last_name="Staff")
        book = Book.objects.create(added_by=other, **self.book_payload())

        response = self.client.post(reverse("library_delete_employee", args=[other.id]))
        book.refresh_from_db()

        self.assertRedirects(response, reverse("library_admin_dashboard"), fetch_redirect_response=False)
        self.assertFalse(Employee.objects.filter(pk=other.id).exists())
        self.assertTrue(Book.objects.filter(pk=book.id).exists())
        self.assertIsNone(book.added_by)

    def test_admin_can_disable_and_enable_employee_without_touching_books(self):
        self.employee.is_admin = True
        self.employee.save(update_fields=["is_admin"])
        other = Employee.objects.create(email="other@example.com", first_name="Other", last_name="Staff")
        other.set_password("StrongPass123!")
        other.save(update_fields=["password_hash"])
        book = Book.objects.create(added_by=other, **self.book_payload())
        self.login()

        response = self.client.post(reverse("library_toggle_employee_status", args=[other.id]))
        other.refresh_from_db()
        book.refresh_from_db()

        self.assertRedirects(response, reverse("library_admin_dashboard"), fetch_redirect_response=False)
        self.assertFalse(other.is_active)
        self.assertEqual(book.added_by, other)

        self.client.logout()
        response = self.client.post(reverse("library_login"), {
            "email": other.email,
            "password": "StrongPass123!",
        })
        self.assertEqual(response.status_code, 200)
        self.assertNotIn("library_employee_id", self.client.session)
        self.assertContains(response, "This account has been disabled")

        self.login()
        response = self.client.post(reverse("library_toggle_employee_status", args=[other.id]))
        other.refresh_from_db()

        self.assertRedirects(response, reverse("library_admin_dashboard"), fetch_redirect_response=False)
        self.assertTrue(other.is_active)

    def test_admin_cannot_delete_self(self):
        self.employee.is_admin = True
        self.employee.save(update_fields=["is_admin"])
        self.login()

        response = self.client.post(reverse("library_delete_employee", args=[self.employee.id]))

        self.assertRedirects(response, reverse("library_admin_dashboard"), fetch_redirect_response=False)
        self.assertTrue(Employee.objects.filter(pk=self.employee.id).exists())
