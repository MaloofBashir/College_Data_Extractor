from functools import wraps
import random

from django.http import HttpResponse
from django.contrib import messages
from django.core.exceptions import ValidationError
from django.db.models import Count, Q, Sum
from django.db.models.functions import Trim
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils.http import url_has_allowed_host_and_scheme
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from .forms import BookForm, EmployeeLoginForm, EmployeeRegistrationForm
from .models import Book, Employee


BULK_BOOK_HEADERS = ["Title", "Author", "Accession Number", "Locker No", "Subject", "Remarks"]
BULK_BOOK_REQUIRED_HEADERS = {"Title", "Author", "Locker No", "Subject"}
BULK_BOOK_FIELD_MAP = {
    "Title": "title",
    "Author": "author",
    "Accession Number": "accession_number",
    "Locker No": "locker_no",
    "Subject": "subject",
    "Remarks": "remarks",
}


def current_employee(request):
    employee_id = request.session.get("library_employee_id")
    if not employee_id:
        return None
    try:
        employee = Employee.objects.get(pk=employee_id)
        if not employee.is_active:
            request.session.pop("library_employee_id", None)
            return None
        return employee
    except Employee.DoesNotExist:
        request.session.pop("library_employee_id", None)
        return None


def require_employee(view_func):
    @wraps(view_func)
    def wrapped(request, *args, **kwargs):
        employee = current_employee(request)
        if not employee:
            return redirect(f"{reverse('library_login')}?next={request.path}")
        request.library_employee = employee
        return view_func(request, *args, **kwargs)

    return wrapped


def require_admin(view_func):
    @wraps(view_func)
    @require_employee
    def wrapped(request, *args, **kwargs):
        if not request.library_employee.is_admin:
            messages.error(request, "Admin access is required.")
            return redirect("library_dashboard")
        return view_func(request, *args, **kwargs)

    return wrapped


def stats_payload(books=None):
    books = books if books is not None else Book.objects.all()
    return {
        "total_books": books.count(),
        "total_quantity": books.aggregate(total=Sum("quantity"))["total"] or 0,
    }


def admin_summary_payload():
    return {
        **stats_payload(),
        "employee_total": Employee.objects.count(),
        "subject_counts": (
            Book.objects.values("subject")
            .annotate(book_count=Count("id"), total_quantity=Sum("quantity"))
            .order_by("subject")
        ),
        "locker_counts": (
            Book.objects.values("locker_no")
            .annotate(book_count=Count("id"), total_quantity=Sum("quantity"))
            .order_by("locker_no")
        ),
        "employee_counts": (
            Employee.objects.annotate(book_count=Count("books"), total_quantity=Sum("books__quantity"))
            .order_by("first_name", "last_name")
        ),
    }


def filter_books(request):
    books = Book.objects.select_related("added_by")
    query = request.GET.get("q", "").strip()
    subject = request.GET.get("subject", "").strip()

    if query:
        books = books.filter(
            Q(title__icontains=query)
            | Q(author__icontains=query)
            | Q(accession_number__icontains=query)
            | Q(locker_no__icontains=query)
            | Q(subject__icontains=query)
        )
    if subject:
        books = books.annotate(clean_subject=Trim("subject")).filter(clean_subject__iexact=subject)

    return books, query, subject


def landing(request):
    employee = current_employee(request)
    if employee:
        return redirect("library_dashboard")
    return render(request, "Library/auth.html", {
        "login_form": EmployeeLoginForm(),
        "registration_form": EmployeeRegistrationForm(),
        "next_url": request.GET.get("next", ""),
    })


def register(request):
    if request.method != "POST":
        return redirect("library_home")
    form = EmployeeRegistrationForm(request.POST)
    if form.is_valid():
        employee = form.save()
        request.session["library_employee_id"] = employee.id
        messages.success(request, "Registration completed. You can start adding books now.")
        return redirect("library_dashboard")
    return render(request, "Library/auth.html", {
        "login_form": EmployeeLoginForm(),
        "registration_form": form,
        "active_panel": "register",
    })


def login_view(request):
    if request.method != "POST":
        return render(request, "Library/auth.html", {
            "login_form": EmployeeLoginForm(),
            "registration_form": EmployeeRegistrationForm(),
            "next_url": request.GET.get("next", ""),
        })
    form = EmployeeLoginForm(request.POST)
    if form.is_valid():
        employee = Employee.objects.filter(email__iexact=form.cleaned_data["email"]).first()
        password = form.cleaned_data["password"]
        if employee and not employee.is_active:
            form.add_error(None, "This account has been disabled. Contact the library admin.")
        elif employee and (employee.check_password(password) or not employee.password_hash):
            if not employee.password_hash:
                employee.set_password(password)
                employee.save(update_fields=["password_hash"])
            request.session["library_employee_id"] = employee.id
            messages.success(request, f"Welcome, {employee.full_name}.")
            next_url = request.POST.get("next", "")
            if next_url and url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}):
                return redirect(next_url)
            return redirect("library_dashboard")
        form.add_error(None, "Email address or password is incorrect.")
    return render(request, "Library/auth.html", {
        "login_form": form,
        "registration_form": EmployeeRegistrationForm(),
        "active_panel": "login",
        "next_url": request.POST.get("next", ""),
    })


def logout_view(request):
    request.session.pop("library_employee_id", None)
    messages.success(request, "You have been logged out.")
    return redirect("library_home")


@require_employee
def dashboard(request):
    employee = request.library_employee
    edit_book = None
    if request.GET.get("edit"):
        edit_book = get_object_or_404(Book, pk=request.GET["edit"], added_by=employee)

    if request.method == "POST":
        edit_id = request.POST.get("book_id")
        instance = get_object_or_404(Book, pk=edit_id, added_by=employee) if edit_id else None
        form = BookForm(request.POST, instance=instance)
        if form.is_valid():
            book = form.save(commit=False)
            book.added_by = employee
            book.save()
            messages.success(request, "Book updated." if instance else "Book added.")
            return redirect("library_dashboard")
    else:
        form = BookForm(instance=edit_book)

    author_suggestions = Book.objects.exclude(author="").values_list("author", flat=True).distinct().order_by("author")
    subjects = Book.objects.exclude(subject="").values_list("subject", flat=True).distinct().order_by("subject")
    own_books = employee.books.select_related("added_by").all()

    return render(request, "Library/dashboard.html", {
        "employee": employee,
        "form": form,
        "edit_book": edit_book,
        "subjects": subjects,
        "author_suggestions": author_suggestions,
        "visible_books": own_books[:10],
        "remaining_books": own_books[10:],
        **stats_payload(employee.books.all()),
    })


@require_employee
def delete_book(request, book_id):
    employee = request.library_employee
    books = Book.objects.all() if employee.is_admin else Book.objects.filter(added_by=employee)
    book = get_object_or_404(books, pk=book_id)
    session_key = f"delete_book_captcha_{book.id}"

    if request.method == "POST":
        expected_answer = request.session.get(session_key)
        entered_answer = request.POST.get("captcha_answer", "").strip()
        if expected_answer is not None and entered_answer == str(expected_answer):
            book.delete()
            request.session.pop(session_key, None)
            messages.success(request, "Book deleted.")
            return redirect("library_admin_dashboard" if employee.is_admin else "library_dashboard")
        messages.error(request, "Captcha answer was incorrect. Please try again.")

    first_number = random.randint(2, 9)
    second_number = random.randint(2, 9)
    request.session[session_key] = first_number + second_number
    return render(request, "Library/delete_book_confirm.html", {
        "employee": employee,
        "book": book,
        "first_number": first_number,
        "second_number": second_number,
    })


def all_books(request):
    query = request.GET.get("q", "").strip()
    selected_subject = request.GET.get("subject", "").strip()
    employee = current_employee(request)
    books = Book.objects.select_related("added_by")
    if employee and not employee.is_admin:
        books = books.filter(added_by=employee)
    subjects = books.exclude(subject="").values_list("subject", flat=True).distinct().order_by("subject")
    return render(request, "Library/all_books.html", {
        "employee": employee,
        "all_books": books,
        "query": query,
        "selected_subject": selected_subject,
        "subjects": subjects,
        "show_all_books": employee is None or employee.is_admin,
        **stats_payload(books),
    })


@require_admin
def admin_dashboard(request):
    return render(request, "Library/admin_dashboard.html", {
        "employee": request.library_employee,
        "all_books": Book.objects.select_related("added_by").all(),
        **admin_summary_payload(),
    })


@require_admin
def delete_employee(request, employee_id):
    employee_to_delete = get_object_or_404(Employee, pk=employee_id)
    if employee_to_delete.id == request.library_employee.id:
        messages.error(request, "You cannot delete your own admin account.")
        return redirect("library_admin_dashboard")

    if request.method == "POST":
        employee_name = employee_to_delete.full_name or employee_to_delete.email
        employee_to_delete.delete()
        messages.success(request, f"{employee_name} was deleted. Their book records were kept.")
        return redirect("library_admin_dashboard")

    return render(request, "Library/delete_employee_confirm.html", {
        "employee": request.library_employee,
        "employee_to_delete": employee_to_delete,
        "book_count": employee_to_delete.books.count(),
    })


@require_admin
def toggle_employee_status(request, employee_id):
    if request.method != "POST":
        return redirect("library_admin_dashboard")

    employee_to_update = get_object_or_404(Employee, pk=employee_id)
    if employee_to_update.id == request.library_employee.id:
        messages.error(request, "You cannot disable your own admin account.")
        return redirect("library_admin_dashboard")

    employee_to_update.is_active = not employee_to_update.is_active
    employee_to_update.save(update_fields=["is_active"])
    state = "enabled" if employee_to_update.is_active else "disabled"
    messages.success(request, f"{employee_to_update.full_name or employee_to_update.email} was {state}. Their book records were not changed.")
    return redirect("library_admin_dashboard")


def add_sheet(workbook, title, headers, rows):
    sheet = workbook.create_sheet(title)
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for row in rows:
        sheet.append(row)
    sheet.freeze_panes = "A2"
    for column in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column) + 2
        sheet.column_dimensions[column[0].column_letter].width = min(width, 42)


def excel_response(workbook, filename):
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response


def normalize_excel_value(value):
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def extract_bulk_book_rows(worksheet):
    header_values = [normalize_excel_value(cell.value) for cell in worksheet[1]]
    header_lookup = {header: index for index, header in enumerate(header_values) if header}
    missing_headers = [header for header in BULK_BOOK_HEADERS if header not in header_lookup]
    if missing_headers:
        raise ValidationError(f"Missing column(s): {', '.join(missing_headers)}.")

    rows = []
    errors = []
    for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        values = {
            header: normalize_excel_value(row[header_lookup[header]]) if header_lookup[header] < len(row) else ""
            for header in BULK_BOOK_HEADERS
        }
        if not any(values.values()):
            continue

        missing_values = [header for header in BULK_BOOK_REQUIRED_HEADERS if not values[header]]
        if missing_values:
            errors.append(f"Row {row_number}: missing {', '.join(missing_values)}.")
            continue

        rows.append({
            "row_number": row_number,
            "data": {field: values[header] for header, field in BULK_BOOK_FIELD_MAP.items()},
        })
    return rows, errors


@require_admin
def export_books_excel(request):
    workbook = Workbook()
    workbook.remove(workbook.active)
    books = Book.objects.select_related("added_by").order_by("subject", "locker_no", "title")
    add_sheet(workbook, "All Books", [
        "Title", "Author", "Accession Number", "Locker No", "Subject",
        "Remarks", "Entered By", "Employee Email", "Created At", "Updated At",
    ], [
        [
            book.title, book.author, book.accession_number, book.locker_no, book.subject,
            book.remarks,
            book.added_by.full_name if book.added_by else "Deleted user",
            book.added_by.email if book.added_by else "",
            book.created_at.strftime("%Y-%m-%d %H:%M"), book.updated_at.strftime("%Y-%m-%d %H:%M"),
        ]
        for book in books
    ])
    summary = admin_summary_payload()
    add_sheet(workbook, "By Subject", ["Subject", "Book Records"], [
        [item["subject"], item["book_count"]] for item in summary["subject_counts"]
    ])
    add_sheet(workbook, "By Locker", ["Locker No", "Book Records"], [
        [item["locker_no"], item["book_count"]] for item in summary["locker_counts"]
    ])
    add_sheet(workbook, "By Employee", ["Employee", "Email", "Book Records"], [
        [item.full_name, item.email, item.book_count] for item in summary["employee_counts"]
    ])
    return excel_response(workbook, "library_books_report.xlsx")


@require_admin
def download_book_upload_template(request):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Books Upload"
    sheet.append(BULK_BOOK_HEADERS)
    sheet.append([
        "Introduction to History",
        "A. Writer",
        "ACC-1001",
        "L-12",
        "History",
        "Core reading",
    ])
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    sheet.freeze_panes = "A2"
    for column in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column) + 2
        sheet.column_dimensions[column[0].column_letter].width = min(width, 42)
    return excel_response(workbook, "library_books_upload_format.xlsx")


@require_admin
def bulk_upload_books(request):
    if request.method != "POST":
        return redirect("library_admin_dashboard")

    upload = request.FILES.get("books_file")
    if not upload:
        messages.error(request, "Choose an Excel file before uploading.")
        return redirect("library_admin_dashboard")

    if not upload.name.lower().endswith(".xlsx"):
        messages.error(request, "Upload a .xlsx Excel file using the provided format.")
        return redirect("library_admin_dashboard")

    try:
        workbook = load_workbook(upload, read_only=True, data_only=True)
        rows, row_errors = extract_bulk_book_rows(workbook.active)
    except Exception as exc:
        messages.error(request, f"Could not read the Excel file. {exc}")
        return redirect("library_admin_dashboard")

    created_count = 0
    for row in rows:
        form = BookForm(row["data"])
        if form.is_valid():
            book = form.save(commit=False)
            book.added_by = request.library_employee
            book.save()
            created_count += 1
        else:
            row_errors.append(f"Row {row['row_number']}: {'; '.join(form.errors.as_text().splitlines())}")

    if created_count:
        messages.success(request, f"{created_count} book record{'s' if created_count != 1 else ''} uploaded.")
    if row_errors:
        preview = " ".join(row_errors[:5])
        remaining = len(row_errors) - 5
        if remaining > 0:
            preview = f"{preview} {remaining} more row{'s' if remaining != 1 else ''} need attention."
        messages.error(request, preview)
    if not created_count and not row_errors:
        messages.error(request, "No book rows were found in the Excel file.")
    return redirect("library_admin_dashboard")
